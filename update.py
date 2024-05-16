import openpyxl
from googleapiclient.discovery import build
from google.oauth2 import service_account
import json
from PySide2.QtWidgets import *



file_path = 'Assests/json/keys.json'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
with open(file_path) as f:
    credentials_info = json.load(f)
# credentials_info = json.loads('Assests/json/keys.json')
creds = service_account.Credentials.from_service_account_info(credentials_info, scopes=SCOPES)
SAMPLE_SPREADSHEET_ID = '1cVvWPTNWRRmxcPkA872LdXikqQoMm5-93lq3MiqYygw'
service = build('sheets', 'v4', credentials=creds, static_discovery=False)

#googlesheet declaration
sheet = service.spreadsheets()

#function for Student details data
def  get_student_details(path,newcombo):
    # Initialize variables to store column indices
    name_column_index = None
    roll_no_column_index = None
    father_name_column_index = None
    whatsapp_column_index = None
    mother_column_index = None
    parent_contact_column_index = None
    address_column_index = None

    # Initialize list to store data for all persons
    persons_data = []
    #to read excel sheet data for Student Details
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    # Iterate through rows to find the "Name" header
    for row in sheet_obj.iter_rows(min_row=1, max_row=3):
        for idx, cell in enumerate(row):
            print(cell.value)
            if cell.value == "Name":
                name_column_index = idx
            elif cell.value == "Roll No":
                roll_no_column_index = idx
            elif cell.value == "Father's Name":
                father_name_column_index = idx
            elif cell.value == "Mother's Name":
                mother_column_index = idx
            elif cell.value == "WhatsApp Number":
                whatsapp_column_index = idx
            elif cell.value == "Father's Mobile No": 
                parent_contact_column_index = idx
            elif cell.value == "Postal Address":
                address_column_index = idx
            



    # If any of the required headers are not found, print an error message and exit
    if None in (name_column_index, roll_no_column_index, father_name_column_index, whatsapp_column_index):
        print("Required headers not found in the Excel file.")
        exit()

    
    # Flag to skip the first row
    first_row_skipped = False

    # Iterate through rows to find all persons and extract their data
    for row in sheet_obj.iter_rows(min_row=2, values_only=True):
        if not first_row_skipped:
          first_row_skipped = True
          continue  # Skip the first row
        
        if row[name_column_index ]:
            person_data = {
                "Name": row[name_column_index ],
                "Roll_No": row[roll_no_column_index ],
                "Father's_Name": row[father_name_column_index ],
                "Mother's_Name":row[mother_column_index],
                "WhatsApp_Number": row[whatsapp_column_index ],
                "Parent_Contact_Number" : row[parent_contact_column_index] ,
                'Address': row[address_column_index ]
            }
            #print(person_data)
            persons_data.append(person_data) 
    
        # Prepare data for updating Google Sheet
    values_to_update = [[person_data["Roll_No"], person_data["Name"], person_data["Father's_Name"], person_data["Mother's_Name"], person_data["WhatsApp_Number"], person_data["Parent_Contact_Number"], person_data["Address"]] for person_data in persons_data]

    # Update Google Sheet with data for all persons
    if values_to_update:
        range_to_update = newcombo+"!A2"  # Example range to update
        request_body = {
            "values": values_to_update
        }
        request = sheet.values().update(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                                        range=range_to_update,
                                        valueInputOption="USER_ENTERED",
                                        body=request_body)
        response = request.execute()
        print("Updated Google Sheet successfully.")
        QMessageBox.warning(None, "Student's Detail Update", "Student's Detail Update Successfully")
    else:
        print("No data to update in Google Sheet.")





#fee details part
def  get_fees(path,newcombo):
    # Load the fee details Excel file
    fee_wb = openpyxl.load_workbook(path)
    fee_sheet = fee_wb.active
    # print(newcombo)
    #Initialize variable for fee
    fee_name_column_index = None
    fee_roll_column_index = None
    otp_no_column_index = None 
    first_inst_name_column_index = None
    second_inst_column_index = None

    #Iterate over the fee details
    for row in fee_sheet.iter_rows(min_row=1, max_row=4):
        for idx, cell in enumerate(row):
            print(cell.value)
            if cell.value == "Name":
                fee_name_column_index = idx
            elif cell.value == "OTP":
                otp_no_column_index = idx
            elif cell.value == "1st Inst.":
                first_inst_name_column_index = idx
            elif cell.value == "2nd inst":
                second_inst_column_index = idx
            elif cell.value == "Roll No.":
                fee_roll_column_index = idx
            elif newcombo[-1] == "I":
                fee_sem_val =1
            elif newcombo[-1:-3] == "II":
                fee_sem_val =2
            elif newcombo[-1:-4] == "III":
                fee_sem_val =3
    
    print(fee_sem_val)
    # print(otp_no_column_index)

    # Initialize list to store data for all persons fee details
    persons_data_fee = []

    # Flag to skip the first row
    first_row_skipped_fee = False

    # Iterate through rows to find all persons and extract their data
    for row in fee_sheet.iter_rows(min_row=2, values_only=True):
        if row[fee_name_column_index]:
            person_data_fee = {
                "Name": row[fee_name_column_index ],
                "Roll No":row[fee_roll_column_index],
                "OTP": row[otp_no_column_index ],
                "1Inst": row[first_inst_name_column_index ],
                "2inst": row[second_inst_column_index ]
            }
            print(person_data_fee)
            last_val = person_data_fee["Roll No"]
            print(last_val)
            persons_data_fee.append(person_data_fee) 

    try:
        last_three_digit=last_val[-3:]
    except:
        last_three_digit="65"

## To read data from google sheet

    # Assuming 'sheet' is already defined and authorized
    # Define the range from which you want to read data
    range_to_read = f""+newcombo+f"!A2:A{last_three_digit}"  # Example range to read, adjust as needed

    # Make a request to get the values from the specified range
    response = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=range_to_read).execute()

    # Get the values from the response
    values = response.get('values', [])
    #print(values)


    #compare against the excel data fee
    google_sheet_roll_numbers = [row[0] for row in response.get('values', [])]
    status_otp=""
    status_first=""
    status_second=""
    # Iterate through fee details from Excel sheet
    for row in fee_sheet.iter_rows(min_row=2, values_only=True):
        fee_roll_number = row[fee_roll_column_index ]  # Assuming fee_roll_column_index is 1-based
        # print(fee_roll_number)
        # print(google_sheet_roll_numbers)
        if fee_roll_number in google_sheet_roll_numbers:
            google_sheet_index = google_sheet_roll_numbers.index(fee_roll_number)
            # Update corresponding fields in Google Sheet
            
            if(row[otp_no_column_index]):
                status_otp="Paid"
            else:
                status_otp=""
            if(row[first_inst_name_column_index]):
                status_first="Paid"
            else:
                status_first=""
            if(row[second_inst_column_index]):
                status_second="Paid"
            else:
                status_second=""

            if(fee_sem_val == 1 or fee_sem_val==0):
                range_to_update = f""+newcombo+f"!N{google_sheet_index+2}"  # Adjust range as needed
            elif(fee_sem_val == 2):
                range_to_update = f""+newcombo+f"!Q{google_sheet_index+2}"  # Adjust range as needed
            elif(fee_sem_val == 3):
                range_to_update = f""+newcombo+f"!T{google_sheet_index+2}"  # Adjust range as needed
            values_to_update = [[status_otp,status_first,status_second]]  # Assuming OTP value is in the first column
            request_body = {"values": values_to_update}
            request = sheet.values().update(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=range_to_update, valueInputOption="USER_ENTERED", body=request_body)
            response = request.execute()
    QMessageBox.warning(None, "Student's Fees Detail Update", "Student's Fees Detail Update Successfully")
            

# for attendance  related data

def get_attendance(path,newcombo,newcombosem):
    print(newcombo)
    # Load the fee details Excel file
    attendance_wb = openpyxl.load_workbook(path,data_only=True)
    attendance_sheet = attendance_wb.active
    # attendance_wb.calculate()

    # Initialize variables to store column indices for attendance sheet
    attendance_roll_column_index = None
    attendance_cum_column_index = None
    attenden_sem_val = None


    # Iterate through rows to find the "Name" header
    for row in attendance_sheet.iter_rows(min_row=1, max_row=6):
        for idx, cell in enumerate(row):
            if cell.value == "Roll No":
                attendance_roll_column_index = idx
            elif cell.value == "Cum_%":
                attendance_cum_column_index = idx
            elif newcombosem == "I":
                attenden_sem_val =1
            elif newcombosem == "II":
                attenden_sem_val =2
            elif newcombosem == "III":
                attenden_sem_val =3
            elif newcombosem == "IV":
                attenden_sem_val =4
            elif newcombosem == "V":
                attenden_sem_val =5
            elif newcombosem == "VI":
                attenden_sem_val =6
                

    print(attendance_roll_column_index)
    print(attendance_cum_column_index)
    # print(attendance_sheet['L9'].value)

    # If any of the required headers are not found, print an error message and exit
    if None in (attendance_cum_column_index, attendance_roll_column_index):
        print("Required headers not found in the Excel file.")
        exit()

    persons_data_attendance = []
    
    for i, row in enumerate(attendance_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[attendance_roll_column_index ]:
            person_data_attendance = {
                "Roll": row[attendance_roll_column_index],
                "Cum": row[attendance_cum_column_index]
            }
            last_val=person_data_attendance["Roll"]
            # print(person_data_attendance)
            persons_data_attendance.append(person_data_attendance)
    print(last_val)
    last_three_digits = last_val[-3:]
        
    # To read data from google sheet

    # Assuming 'sheet' is already defined and authorized
    # Define the range from which you want to read data
    range_to_read = newcombo+f"!A2:A{last_three_digits}"  # Example range to read, adjust as needed

    # Make a request to get the values from the specified range
    response = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=range_to_read).execute()

    # Get the values from the response
    values = response.get('values', [])
    #print(values)


    #compare against the excel data fee
    google_sheet_names = [row[0] for row in response.get('values', [])]
    # print(google_sheet_names)
    # Iterate through fee details from Excel sheet
    for row in attendance_sheet.iter_rows(min_row=2, values_only=True):
        # fee_roll_number = "BCA202200" + str(row[fee_roll_column_index ])  # Assuming fee_roll_column_index is 1-based
        attendance_roll = row[attendance_roll_column_index]
        #print(attendance_name)
        if attendance_roll in google_sheet_names:
            google_sheet_index = google_sheet_names.index(attendance_roll)
            # print(google_sheet_index)
            # Update corresponding fields in Google Sheet
            if(attenden_sem_val == 1):
                range_to_update = newcombo+f"!H{google_sheet_index+2}"  # Adjust range as needed
            elif(attenden_sem_val == 2):
                range_to_update = newcombo+f"!I{google_sheet_index+2}"    
            elif(attenden_sem_val == 3):
                range_to_update = newcombo+f"!J{google_sheet_index+2}"   
            elif(attenden_sem_val == 4):
                range_to_update = newcombo+f"!K{google_sheet_index+2}"
            elif(attenden_sem_val == 5):
                range_to_update = newcombo+f"!L{google_sheet_index+2}"
            elif(attenden_sem_val == 6):
                range_to_update = newcombo+f"!M{google_sheet_index+2}" 
            values_to_update = [[row[attendance_cum_column_index ]]]  # Assuming OTP value is in the first column
            request_body = {"values": values_to_update}
            request = sheet.values().update(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=range_to_update, valueInputOption="USER_ENTERED", body=request_body)
            response = request.execute()
    QMessageBox.warning(None, "Student's Attendance Detail Update", "Student's Attendance Detail Update Successfully")
            # print(f"Updated data in Google Sheet for Roll No: {attendance_roll}.")
        # else:
        #     print("Not updated")



# function to get marks details from excel sheet and update it
            
def get_internal_marks(path,newcombo,newcombosem):
        # Load the fee details Excel file
    marks_wb = openpyxl.load_workbook(path)
    marks_sheet = marks_wb.active

    # Initialize variables to store column indices for attendance sheet
    marks_roll_column_index = None
    marks_internal_column_index = None
    
    marks_sem_val = None

    # Iterate through rows to find the "Name" header
    for row in marks_sheet.iter_rows(min_row=1, max_row=10):
        for idx, cell in enumerate(row):
            if cell.value == "Roll No":
                marks_roll_column_index = idx
            elif cell.value == "SCGPA":
                marks_internal_column_index= idx
            elif newcombosem == "I":
                marks_sem_val=1
            elif newcombosem == "II":
                marks_sem_val=2
            elif newcombosem == "III":
                marks_sem_val=3
            elif newcombosem == "IV":
                marks_sem_val=4
            elif newcombosem == "V":
                marks_sem_val=5
            elif newcombosem == "VI":
                marks_sem_val=6
    # print(marks_roll_column_index)
    # print(marks_internal_column_index)



    # If any of the required headers are not found, print an error message and exit
    if None in (marks_roll_column_index, marks_internal_column_index ):
        print("Required headers not found in the Excel file.")
        exit()

    persons_data_marks =[]

    for row in marks_sheet.iter_rows(min_row=2, values_only=True):
        #if not first_row_skipped:
        #   first_row_skipped = True
        #   continue  # Skip the first row
        if row[marks_roll_column_index ]:
            person_data_marks = {
                "Roll": row[marks_roll_column_index ],
                "Internal":row[marks_internal_column_index],
        

            }
            print(person_data_marks)
            last_val=person_data_marks["Roll"]
            print(last_val)
            persons_data_marks.append(person_data_marks)

    last_three_digit = last_val[-3:]
    print(last_three_digit)
    # Define the range from which you want to read data
    range_to_read = newcombo+f"!A2:A{last_three_digit}"  # Example range to read, adjust as needed

    # Make a request to get the values from the specified range
    response = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=range_to_read).execute()

    # Get the values from the response
    values = response.get('values', [])
    #print(values)


    #compare against the excel data fee
    google_sheet_names = [row[0] for row in response.get('values', [])]

    
    # Iterate through fee details from Excel sheet
    for row in marks_sheet.iter_rows(min_row=2, values_only=True):
        #fee_roll_number = "BCA202200" + str(row[fee_roll_column_index ])  # Assuming fee_roll_column_index is 1-based
        marks_roll = row[marks_roll_column_index]
        #print(attendance_name)
        if marks_roll in google_sheet_names:
            google_sheet_index = google_sheet_names.index(marks_roll)
            # Update corresponding fields in Google Sheet
            if(marks_sem_val==1 or marks_sem_val==None):
                range_to_update = newcombo+f"!W{google_sheet_index+2}"  # Adjust range as needed
            elif(marks_sem_val==2):
                range_to_update = newcombo+f"!X{google_sheet_index+2}"  # Adjust range as needed
            elif(marks_sem_val==3):
                range_to_update = newcombo+f"!Y{google_sheet_index+2}"  # Adjust range as needed
            elif(marks_sem_val==4):
                range_to_update = newcombo+f"!Z{google_sheet_index+2}"  # Adjust range as needed
            elif(marks_sem_val==5):
                range_to_update = newcombo+f"!AA{google_sheet_index+2}"  # Adjust range as needed
            elif(marks_sem_val==6):
                range_to_update = newcombo+f"!AB{google_sheet_index+2}"  # Adjust range as needed
            values_to_update = [[row[marks_internal_column_index ]]]  # Assuming OTP value is in the first column
            request_body = {"values": values_to_update}
            request = sheet.values().update(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=range_to_update, valueInputOption="USER_ENTERED", body=request_body)
            response = request.execute()
    QMessageBox.warning(None, "Student's Internal Marks Update", "Student's Internal Marks Update Successfully")
        #     print(f"Updated data in Google Sheet for Roll No: {marks_roll}.")
        # else:
        #     print("Not updated")


def get_external_marks(path,newcombo,newcombosem):
        # Load the fee details Excel file
    marks_wb = openpyxl.load_workbook(path)
    marks_sheet = marks_wb.active

    # Initialize variables to store column indices for attendance sheet
    marks_roll_column_index = None
    marks_external_column_index = None
    
    marks_sem_val = None

    # Iterate through rows to find the "Name" header
    for row in marks_sheet.iter_rows(min_row=1, max_row=10):
        for idx, cell in enumerate(row):
            if cell.value == "Roll No":
                marks_roll_column_index = idx
            elif cell.value == "SCGPA":
                marks_external_column_index= idx
            elif newcombosem == "I":
                marks_sem_val=1
            elif newcombosem == "II":
                marks_sem_val=2
            elif newcombosem == "III":
                marks_sem_val=3
            elif newcombosem == "IV":
                marks_sem_val=4
            elif newcombosem == "V":
                marks_sem_val=5
            elif newcombosem == "VI":
                marks_sem_val=6
    



    # If any of the required headers are not found, print an error message and exit
    if None in (marks_roll_column_index, marks_external_column_index ):
        print("Required headers not found in the Excel file.")
        exit()

    persons_data_marks =[]

    for row in marks_sheet.iter_rows(min_row=2, values_only=True):
        #if not first_row_skipped:
        #   first_row_skipped = True
        #   continue  # Skip the first row
        if row[marks_roll_column_index ]:
            person_data_marks = {
                "Roll": row[marks_roll_column_index ],
                "External":row[marks_external_column_index],
        

            }
            print(person_data_marks)
            last_val=person_data_marks["Roll"]
            print(last_val)
            persons_data_marks.append(person_data_marks)

    last_three_digit = last_val[-3:]
    print(last_three_digit)
    # Define the range from which you want to read data
    range_to_read = newcombo+f"!A2:A{last_three_digit}"  # Example range to read, adjust as needed

    # Make a request to get the values from the specified range
    response = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=range_to_read).execute()

    # Get the values from the response
    values = response.get('values', [])
    #print(values)


    #compare against the excel data fee
    google_sheet_names = [row[0] for row in response.get('values', [])]

    
    # Iterate through fee details from Excel sheet
    for row in marks_sheet.iter_rows(min_row=2, values_only=True):
        #fee_roll_number = "BCA202200" + str(row[fee_roll_column_index ])  # Assuming fee_roll_column_index is 1-based
        marks_roll = row[marks_roll_column_index]
        #print(attendance_name)
        if marks_roll in google_sheet_names:
            google_sheet_index = google_sheet_names.index(marks_roll)
            # Update corresponding fields in Google Sheet
            if(marks_sem_val==1 or  marks_sem_val==None):
                range_to_update = newcombo+f"!AC{google_sheet_index+2}"  # Adjust range as needed
            elif(marks_sem_val==2):
                range_to_update = newcombo+f"!AD{google_sheet_index+2}"  # Adjust range as needed
            elif(marks_sem_val==3):
                range_to_update = newcombo+f"!AE{google_sheet_index+2}"  # Adjust range as needed
            elif(marks_sem_val==4):
                range_to_update = newcombo+f"!AF{google_sheet_index+2}"  # Adjust range as needed
            elif(marks_sem_val==5):
                range_to_update = newcombo+f"!AG{google_sheet_index+2}"  # Adjust range as needed
            elif(marks_sem_val==6):
                range_to_update = newcombo+f"!AH{google_sheet_index+2}"  # Adjust range as needed
            values_to_update = [[row[marks_external_column_index ]]]  # Assuming OTP value is in the first column
            request_body = {"values": values_to_update}
            request = sheet.values().update(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=range_to_update, valueInputOption="USER_ENTERED", body=request_body)
            response = request.execute()
    QMessageBox.warning(None, "Student's External Marks Update", "Student's External Marks Update Successfully")
        
